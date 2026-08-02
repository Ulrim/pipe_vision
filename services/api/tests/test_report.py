"""월간 품질 리포트 생성 검증 (M12). PDF/XLSX 바이트 + 헤더 시그니처."""
from __future__ import annotations

from datetime import datetime, timezone


def _seed_item(client, auth):
    client.post("/master/items", headers=auth("qa1"), json={
        "item_code": "RPT", "item_name": "rpt", "ref_length_mm": 100.0,
        "tol_plus_mm": 0.5, "tol_minus_mm": 0.5, "px_to_mm_scale": 0.05,
    })


def _post(client, **over):
    base = {
        "lot": "R", "item_code": "RPT", "cam_id": "C",
        "inspected_at": datetime(2026, 3, 10, 9, 0, tzinfo=timezone.utc).isoformat(),
        "final_verdict": "OK", "defect_codes": [], "review_flag": False,
        "mes_synced": True, "proc_time_ms": 100,
    }
    base.update(over)
    r = client.post("/inspection", json=base)
    assert r.status_code == 201, r.text


def _seed_month(client, auth):
    _seed_item(client, auth)
    for i in range(5):
        _post(client, lot=f"OK{i}")
    _post(client, lot="NG0", final_verdict="NG", defect_codes=["LEN"], mes_synced=False)
    _post(client, lot="NG1", final_verdict="NG", defect_codes=["OIL", "SCR"],
          mes_synced=False,
          inspected_at=datetime(2026, 3, 11, 9, 0, tzinfo=timezone.utc).isoformat())


def test_report_pdf_signature(client, auth):
    _seed_month(client, auth)
    r = client.get("/kpi/report", headers=auth("qa1"),
                   params={"period": "2026-03", "fmt": "pdf"})
    assert r.status_code == 200, r.text
    assert r.headers["content-type"] == "application/pdf"
    cd = r.headers["content-disposition"]
    assert "attachment" in cd
    assert 'filename="aivis_kpi_2026-03.pdf"' in cd
    # PDF 매직넘버 %PDF-
    assert r.content[:5] == b"%PDF-"
    assert len(r.content) > 500


def test_report_xlsx_signature(client, auth):
    _seed_month(client, auth)
    r = client.get("/kpi/report", headers=auth("qa1"),
                   params={"period": "2026-03", "fmt": "xlsx"})
    assert r.status_code == 200, r.text
    assert r.headers["content-type"] == (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    cd = r.headers["content-disposition"]
    assert "attachment" in cd
    assert 'filename="aivis_kpi_2026-03.xlsx"' in cd
    # XLSX = zip 컨테이너 -> PK\x03\x04
    assert r.content[:4] == b"PK\x03\x04"
    assert len(r.content) > 500


def test_report_requires_quality(client, auth):
    # 작업자는 리포트 생성 불가(403).
    r = client.get("/kpi/report", headers=auth("op1"), params={"period": "2026-03"})
    assert r.status_code == 403


def test_report_default_period_is_current_month(client, auth):
    # period 미지정 -> 당월. 빈 달이어도 정상 생성(분모 0 보호).
    r = client.get("/kpi/report", headers=auth("qa1"), params={"fmt": "xlsx"})
    assert r.status_code == 200
    now = datetime.now(timezone.utc)
    fname = f"aivis_kpi_{now.year:04d}-{now.month:02d}.xlsx"
    assert fname in r.headers["content-disposition"]
    assert r.content[:4] == b"PK\x03\x04"


def test_report_no_auth_rejected(client):
    r = client.get("/kpi/report", params={"period": "2026-03"})
    assert r.status_code == 401


# --- KPI 목표 대비 달성 판정 / 처리속도 백분위 (M12, §1.1·§1.2) --------------
class _Row:
    """집계 함수 입력용 최소 스텁(Inspection 대체)."""

    def __init__(self, proc_time_ms=None, final_verdict="OK", defect_codes=None,
                 inspected_at=None):
        self.proc_time_ms = proc_time_ms
        self.final_verdict = final_verdict
        self.defect_codes = defect_codes or []
        self.inspected_at = inspected_at or datetime(2026, 3, 10, tzinfo=timezone.utc)


def _summary(**over):
    from aivis_types import KpiSummary

    base = dict(
        period="2026-03", total_inspected=100, defect_count=1,
        process_defect_ppm=100.0, auto_inspected=100,
        auto_inspection_rate_pct=100.0, misjudge_count=0, miss_count=0,
        inspection_defect_rate_pct=0.0, stored_count=100, mes_synced_count=100,
        storage_mes_rate_pct=100.0, avg_proc_time_ms=120.0,
    )
    base.update(over)
    return KpiSummary(**base)


def test_proc_time_percentiles():
    from core.report import proc_time_percentiles

    rows = [_Row(proc_time_ms=v) for v in range(1, 101)]  # 1..100ms
    pct = proc_time_percentiles(rows)
    assert pct["p50"] == 50
    assert pct["p95"] == 95
    assert pct["p99"] == 99
    # 표본 없음 → None(판정 보류).
    assert proc_time_percentiles([])["p95"] is None


def test_evaluate_targets_pass_and_fail():
    """§1.1/§1.2 목표 판정: 상한(lte)·하한(gte) 규칙이 정확해야 한다."""
    from core.report import evaluate_targets

    rows = [_Row(proc_time_ms=100) for _ in range(50)]  # p95=100ms → 300 이하 달성
    good = dict(evaluate_targets(_summary(process_defect_ppm=500.0), rows)and
                {k: p for k, _ko, _l, _t, p in
                 evaluate_targets(_summary(process_defect_ppm=500.0), rows)})
    assert good["process_defect_ppm"] is True       # 500 <= 600
    assert good["auto_inspection_rate_pct"] is True  # 100 >= 100
    assert good["p95_proc_time_ms"] is True          # 100 <= 300

    bad = {k: p for k, _ko, _l, _t, p in
           evaluate_targets(
               _summary(process_defect_ppm=700.0, auto_inspection_rate_pct=99.0),
               [_Row(proc_time_ms=400) for _ in range(50)])}
    assert bad["process_defect_ppm"] is False        # 700 > 600
    assert bad["auto_inspection_rate_pct"] is False  # 99 < 100
    assert bad["p95_proc_time_ms"] is False          # 400 > 300


def test_evaluate_targets_none_when_no_proc_data():
    """처리속도 표본이 없으면 판정 보류(None) — 임의로 합격 처리하지 않는다."""
    from core.report import evaluate_targets

    res = {k: p for k, _ko, _l, _t, p in evaluate_targets(_summary(), [])}
    assert res["p95_proc_time_ms"] is None


def test_status_markers_are_ascii_only():
    """달성/미달 기호는 ASCII 여야 한다(색 단독 표기 방지 회귀).

    한글 CID 폰트에 ✓/✗(U+2713/2717) 글리프가 없어 조용히 빈칸으로 렌더링되면
    '달성 여부'가 색으로만 구분되어 색각 이상 사용자가 판별할 수 없게 된다.
    """
    import re
    from pathlib import Path

    src = Path(__file__).resolve().parents[1] / "core" / "report.py"
    text = src.read_text(encoding="utf-8")
    # 상태 표기를 만드는 라인에 비-ASCII 기호가 섞이지 않았는지 확인.
    for m in re.finditer(r'mark, ink = f"(.+?) \{_lab', text):
        assert m.group(1).isascii(), f"비-ASCII 상태 기호: {m.group(1)!r}"


def test_report_pdf_includes_charts_and_targets(client, auth):
    """PDF 에 차트(추세/불량유형)와 목표 판정표가 포함되어 커진다."""
    _seed_month(client, auth)
    r = client.get("/kpi/report", headers=auth("qa1"),
                   params={"period": "2026-03", "fmt": "pdf"})
    assert r.status_code == 200
    assert r.content[:4] == b"%PDF"
    # 표만 있던 시절(약 5KB)보다 확실히 큼 — 차트 벡터가 들어갔다.
    assert len(r.content) > 6000


def test_report_xlsx_has_target_rows(client, auth):
    """XLSX KPI 시트에 목표 대비 달성 행이 있어야 한다(PDF 와 동일 기준)."""
    import io

    from openpyxl import load_workbook

    _seed_month(client, auth)
    r = client.get("/kpi/report", headers=auth("qa1"),
                   params={"period": "2026-03", "fmt": "xlsx"})
    assert r.status_code == 200
    ws = load_workbook(io.BytesIO(r.content))["KPI"]
    cells = [c for row in ws.iter_rows(values_only=True) for c in row if c]
    joined = " ".join(str(c) for c in cells)
    assert "목표 대비 달성" in joined
    assert "공정불량률 (ppm)" in joined
    assert ("달성" in joined) or ("미달" in joined)
