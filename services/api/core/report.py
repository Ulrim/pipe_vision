"""월간 품질 리포트 생성 (CLAUDE.md §5 M12, §1.1).

KpiSummary(§1.1 산출식) + 불량유형별 집계 + 일자별 검사/불량 표를
PDF(reportlab) / XLSX(openpyxl) 바이트로 렌더링한다.

한글 라벨을 기본 사용하되, PDF 에서 한글 폰트 등록에 실패하면
라틴 라벨 + 코드 폴백으로 깨짐 없이 출력한다(M12 요구).
"""
from __future__ import annotations

import io
import math
import os
from collections import Counter, defaultdict
from datetime import datetime, timezone
from typing import Optional

from aivis_types import KpiSummary
from db.models import Inspection

# ---- 라벨(한글/라틴 폴백) -------------------------------------------------

# (한글, 라틴 폴백) 쌍. 한글 폰트 미등록 시 라틴 사용.
LABELS: dict[str, tuple[str, str]] = {
    "title": ("AIVIS 월간 품질 리포트", "AIVIS Monthly Quality Report"),
    "period": ("대상 기간", "Period"),
    "generated": ("생성 시각", "Generated"),
    "kpi": ("KPI 요약 (사업계획서 §1.1)", "KPI Summary (sec 1.1)"),
    "total_inspected": ("총 검사수량", "Total inspected"),
    "process_defect_ppm": ("공정불량률 (ppm)", "Process defect (ppm)"),
    "inspection_defect_rate_pct": ("검사불량률 (%)", "Inspection defect (%)"),
    "auto_inspection_rate_pct": ("자동검사율 (%)", "Auto inspection (%)"),
    "storage_mes_rate_pct": ("저장·MES 연계율 (%)", "Storage/MES link (%)"),
    "avg_proc_time_ms": ("평균 처리속도 (ms)", "Avg proc time (ms)"),
    "defect_count": ("공정 불량수량", "Defect count"),
    "defect_breakdown": ("불량유형별 집계", "Defect breakdown"),
    "code": ("코드", "Code"),
    "count": ("건수", "Count"),
    "daily": ("일자별 검사/불량", "Daily inspected / defects"),
    "date": ("일자", "Date"),
    "inspected": ("검사수", "Inspected"),
    "defects": ("불량수", "Defects"),
    "none": ("해당 없음", "None"),
    "targets": ("KPI 목표 대비 달성 (인수 기준)", "KPI vs Target (acceptance)"),
    "kpi_item": ("항목", "Item"),
    "target": ("목표", "Target"),
    "actual": ("실적", "Actual"),
    "achieved": ("달성 여부", "Achieved"),
    "pass_ko": ("달성", "PASS"),
    "fail_ko": ("미달", "FAIL"),
    "na": ("판정보류", "N/A"),
    "trend": ("일자별 공정불량률 추세 (ppm)", "Daily process defect trend (ppm)"),
    "defect_chart": ("불량유형별 건수", "Defect count by type"),
    "proc_pct": ("처리속도 분포 (ms)", "Proc time distribution (ms)"),
    "p50": ("p50 (중앙값)", "p50 (median)"),
    "p95": ("p95", "p95"),
    "p99": ("p99", "p99"),
}

# 불량유형 코드 한글 설명(§7.2).
DEFECT_KO = {
    "LEN": "길이",
    "OIL": "유분기",
    "DIS": "변색",
    "SCR": "스크래치",
    "MULTI": "복합",
}

# ---- 차트 색(데이터 시각화 규칙) ------------------------------------------
# 단일 시리즈 = 카테고리 슬롯 1 하나만 사용(명목 카테고리에 값-램프 금지).
# 상태색은 예약색이며 **반드시 기호+문자와 함께** 쓴다(적/녹은 색각 이상에서
# 구분되지 않으므로 색 단독 표기 금지).
SERIES_1 = "#2a78d6"
STATUS_GOOD = "#0ca30c"
STATUS_CRITICAL = "#d03b3b"
AXIS_INK = "#52514e"
GRID_INK = "#d8d7d3"

# ---- §1.1/§1.2 목표치(인수 합격 기준) --------------------------------------
# (키, 한글라벨, 라틴라벨, 목표문구, 판정함수) — 판정함수는 실적값→합격여부.
# 목표를 코드 한 곳에 두어 리포트/대시보드가 같은 기준을 쓰게 한다.
KPI_TARGETS: list[tuple[str, str, str, str, str]] = [
    ("process_defect_ppm", "공정불량률 (ppm)", "Process defect (ppm)", "600 이하", "lte:600"),
    ("inspection_defect_rate_pct", "검사불량률 (%)", "Inspection defect (%)", "30 이하", "lte:30"),
    ("auto_inspection_rate_pct", "자동검사율 (%)", "Auto inspection (%)", "100", "gte:100"),
    ("storage_mes_rate_pct", "저장·MES 연계율 (%)", "Storage/MES link (%)", "100", "gte:100"),
    ("p95_proc_time_ms", "처리속도 p95 (ms)", "Proc time p95 (ms)", "300 이하", "lte:300"),
]


def proc_time_percentiles(rows: list[Inspection]) -> dict[str, Optional[float]]:
    """처리속도 백분위(p50/p95/p99). 표본 없으면 None.

    FAT/SAT 기준은 300ms/ea 이므로 평균만으로는 불충분하다(꼬리가 기준을
    넘는지 봐야 한다). 선형보간 없이 최근접 순위법(결정적).
    """
    vals = sorted(int(r.proc_time_ms) for r in rows if r.proc_time_ms is not None)
    if not vals:
        return {"p50": None, "p95": None, "p99": None}

    def _pct(p: float) -> float:
        # 최근접 순위법: rank = ceil(p/100 × N), 값 = 정렬표본[rank-1].
        rank = math.ceil(p / 100.0 * len(vals))
        idx = min(len(vals) - 1, max(0, rank - 1))
        return float(vals[idx])

    return {"p50": _pct(50), "p95": _pct(95), "p99": _pct(99)}


def evaluate_targets(
    summary: KpiSummary, rows: list[Inspection]
) -> list[tuple[str, str, str, str, Optional[bool]]]:
    """§1.1/§1.2 목표 대비 실적 판정.

    반환: (키, 한글라벨, 라틴라벨, 목표문구, 달성여부) 리스트.
    달성여부 None = 실적 데이터 없음(판정 보류). 인수 심사에 그대로 쓰도록
    목표·실적·달성여부를 한 표로 낼 수 있게 한다(§12 인수 산출물).
    """
    pct = proc_time_percentiles(rows)
    actuals: dict[str, Optional[float]] = {
        "process_defect_ppm": summary.process_defect_ppm,
        "inspection_defect_rate_pct": summary.inspection_defect_rate_pct,
        "auto_inspection_rate_pct": summary.auto_inspection_rate_pct,
        "storage_mes_rate_pct": summary.storage_mes_rate_pct,
        "p95_proc_time_ms": pct["p95"],
    }
    out: list[tuple[str, str, str, str, Optional[bool]]] = []
    for key, ko, latin, target_text, rule in KPI_TARGETS:
        val = actuals.get(key)
        if val is None:
            out.append((key, ko, latin, target_text, None))
            continue
        op, bound = rule.split(":")
        limit = float(bound)
        passed = (val <= limit) if op == "lte" else (val >= limit)
        out.append((key, ko, latin, target_text, bool(passed)))
    return out


def target_actual_text(
    key: str, summary: KpiSummary, rows: list[Inspection]
) -> str:
    """목표 판정표에 쓸 실적값 문자열."""
    pct = proc_time_percentiles(rows)
    if key == "p95_proc_time_ms":
        return "-" if pct["p95"] is None else f"{pct['p95']:.0f}"
    val = getattr(summary, key, None)
    return "-" if val is None else f"{float(val):.3f}"


def _lab(key: str, korean_ok: bool) -> str:
    ko, latin = LABELS[key]
    return ko if korean_ok else latin


def _defect_label(code: str, korean_ok: bool) -> str:
    if korean_ok and code in DEFECT_KO:
        return f"{code} ({DEFECT_KO[code]})"
    return code


def defect_label(code: str) -> str:
    """불량유형 코드의 한글 표기(외부 공개용 — 미리보기 API 등)."""
    return _defect_label(code, True)


# ---- 데이터 집계 ----------------------------------------------------------


def aggregate_defects(rows: list[Inspection]) -> list[tuple[str, int]]:
    """defect_codes 배열을 코드별로 집계. 코드 정렬된 (code, count) 리스트."""
    counter: Counter[str] = Counter()
    for r in rows:
        for code in r.defect_codes or []:
            counter[str(code)] += 1
    return sorted(counter.items(), key=lambda x: x[0])


def aggregate_daily(rows: list[Inspection]) -> list[tuple[str, int, int]]:
    """일자(YYYY-MM-DD)별 (검사수, 불량수) 집계. 일자 오름차순."""
    inspected: dict[str, int] = defaultdict(int)
    defects: dict[str, int] = defaultdict(int)
    for r in rows:
        day = r.inspected_at.date().isoformat()
        inspected[day] += 1
        if r.final_verdict == "NG":
            defects[day] += 1
    return [(d, inspected[d], defects.get(d, 0)) for d in sorted(inspected)]


# ---- 차트(reportlab.graphics — 추가 의존성 없음, 파이4 고려) ---------------


def _defect_bar_chart(breakdown: list[tuple[str, int]], font: str, korean_ok: bool):
    """불량유형별 건수 막대차트. 명목 카테고리이므로 **모든 막대 단일 색**.

    (값이 큰 막대를 진하게 칠하는 값-램프는 막대 길이가 이미 보여주는 정보를
    색으로 중복 인코딩하므로 쓰지 않는다.) 각 막대에 직접 값 라벨을 단다.
    """
    from reportlab.graphics.charts.barcharts import VerticalBarChart
    from reportlab.graphics.shapes import Drawing, String
    from reportlab.lib import colors

    d = Drawing(430, 165)
    if not breakdown:
        d.add(String(0, 80, _lab("none", korean_ok), fontName=font, fontSize=10))
        return d
    bc = VerticalBarChart()
    bc.x, bc.y, bc.width, bc.height = 35, 32, 375, 112
    bc.data = [[n for _c, n in breakdown]]
    bc.categoryAxis.categoryNames = [
        _defect_label(c, korean_ok) for c, _n in breakdown
    ]
    bc.bars[0].fillColor = colors.HexColor(SERIES_1)
    bc.bars[0].strokeWidth = 0
    bc.barSpacing = 3
    bc.groupSpacing = 12
    bc.valueAxis.valueMin = 0
    bc.valueAxis.valueMax = max(n for _c, n in breakdown) * 1.18 or 1
    # 축/그리드는 뒤로 물린다(데이터가 주인공).
    bc.valueAxis.strokeColor = colors.HexColor(GRID_INK)
    bc.valueAxis.gridStrokeColor = colors.HexColor(GRID_INK)
    bc.valueAxis.visibleGrid = True
    bc.valueAxis.labels.fontName = font
    bc.valueAxis.labels.fontSize = 8
    bc.valueAxis.labels.fillColor = colors.HexColor(AXIS_INK)
    bc.categoryAxis.strokeColor = colors.HexColor(GRID_INK)
    bc.categoryAxis.labels.fontName = font
    bc.categoryAxis.labels.fontSize = 8
    bc.categoryAxis.labels.fillColor = colors.HexColor(AXIS_INK)
    # 직접 값 라벨(문자는 텍스트 잉크로 — 시리즈 색을 글자에 쓰지 않는다).
    bc.barLabels.fontName = font
    bc.barLabels.fontSize = 8
    bc.barLabels.fillColor = colors.HexColor(AXIS_INK)
    bc.barLabelFormat = "%d"
    bc.barLabels.dy = 6
    d.add(bc)
    return d


def _daily_trend_chart(
    daily: list[tuple[str, int, int]], font: str, korean_ok: bool
):
    """일자별 불량률(ppm) 추세 라인차트 + 목표선(600ppm).

    단일 시리즈이므로 범례 없이 제목이 계열을 지칭한다(§규칙). 목표선은
    점선 + 문자 라벨을 함께 달아 색 없이도 의미가 전달되게 한다.
    """
    from reportlab.graphics.charts.linecharts import HorizontalLineChart
    from reportlab.graphics.shapes import Drawing, Line, String
    from reportlab.lib import colors

    d = Drawing(430, 175)
    if not daily:
        d.add(String(0, 85, _lab("none", korean_ok), fontName=font, fontSize=10))
        return d
    ppm = [
        (defects / inspected * 1_000_000.0) if inspected else 0.0
        for _day, inspected, defects in daily
    ]
    lc = HorizontalLineChart()
    lc.x, lc.y, lc.width, lc.height = 42, 30, 368, 120
    lc.data = [ppm]
    lc.lines[0].strokeColor = colors.HexColor(SERIES_1)
    lc.lines[0].strokeWidth = 2
    lc.joinedLines = 1
    top = max(max(ppm), 600.0) * 1.2
    lc.valueAxis.valueMin = 0
    lc.valueAxis.valueMax = top
    lc.valueAxis.strokeColor = colors.HexColor(GRID_INK)
    lc.valueAxis.gridStrokeColor = colors.HexColor(GRID_INK)
    lc.valueAxis.visibleGrid = True
    lc.valueAxis.labels.fontName = font
    lc.valueAxis.labels.fontSize = 8
    lc.valueAxis.labels.fillColor = colors.HexColor(AXIS_INK)
    # 일자 라벨은 과밀하므로 5일 간격만 표기(라벨 충돌 방지).
    names = []
    for i, (day, _ins, _dfx) in enumerate(daily):
        names.append(day[8:] if (i % 5 == 0 or i == len(daily) - 1) else "")
    lc.categoryAxis.categoryNames = names
    lc.categoryAxis.strokeColor = colors.HexColor(GRID_INK)
    lc.categoryAxis.labels.fontName = font
    lc.categoryAxis.labels.fontSize = 7
    lc.categoryAxis.labels.fillColor = colors.HexColor(AXIS_INK)
    d.add(lc)
    # 목표선(600ppm) — 점선 + 문자 라벨(색 단독 의존 금지).
    y = lc.y + (600.0 / top) * lc.height if top > 0 else lc.y
    if lc.y <= y <= lc.y + lc.height:
        ln = Line(lc.x, y, lc.x + lc.width, y)
        ln.strokeColor = colors.HexColor(STATUS_CRITICAL)
        ln.strokeDashArray = [3, 2]
        ln.strokeWidth = 1
        d.add(ln)
        tgt = "목표 600ppm" if korean_ok else "target 600ppm"
        s = String(lc.x + 2, y + 3, tgt, fontName=font, fontSize=7)
        s.fillColor = colors.HexColor(AXIS_INK)
        d.add(s)
    return d


# ---- PDF ------------------------------------------------------------------


def _register_korean_font() -> Optional[str]:
    """가용한 한글 폰트(TTF)를 reportlab 에 등록. 성공 시 폰트명, 실패 시 None.

    CID 폰트(HYSMyeongJo-Medium) 우선 시도 후 시스템 TTF 탐색.
    """
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.cidfonts import UnicodeCIDFont
    from reportlab.pdfbase.ttfonts import TTFont

    # 1) reportlab 내장 CID 한글 폰트(추가 파일 불필요).
    try:
        pdfmetrics.registerFont(UnicodeCIDFont("HYSMyeongJo-Medium"))
        return "HYSMyeongJo-Medium"
    except Exception:
        pass

    # 2) 시스템 TTF 폴백.
    candidates = [
        "/usr/share/fonts/truetype/nanum/NanumGothic.ttf",
        "/usr/share/fonts/truetype/noto/NotoSansCJK-Regular.ttc",
        "/usr/share/fonts/opentype/noto/NotoSansCJK-Regular.ttc",
        "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",  # 한글 X, 등록만
    ]
    for path in candidates:
        if os.path.exists(path) and path.endswith((".ttf",)):
            try:
                pdfmetrics.registerFont(TTFont("AIVIS-KR", path))
                # NanumGothic 만 한글 지원. DejaVu 는 라틴 폴백 유도.
                if "Nanum" in path or "Noto" in path:
                    return "AIVIS-KR"
            except Exception:
                continue
    return None


def render_pdf(summary: KpiSummary, rows: list[Inspection]) -> bytes:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.platypus import (
        Paragraph,
        SimpleDocTemplate,
        Spacer,
        Table,
        TableStyle,
    )

    kr_font = _register_korean_font()
    korean_ok = kr_font is not None
    font = kr_font or "Helvetica"
    font_bold = kr_font or "Helvetica-Bold"

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=A4,
        leftMargin=18 * mm, rightMargin=18 * mm,
        topMargin=18 * mm, bottomMargin=18 * mm,
        title=f"AIVIS KPI {summary.period}",
    )
    styles = getSampleStyleSheet()
    h1 = styles["Title"]
    h1.fontName = font_bold
    h2 = styles["Heading2"]
    h2.fontName = font_bold
    body = styles["Normal"]
    body.fontName = font

    story = []
    story.append(Paragraph(_lab("title", korean_ok), h1))
    now = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
    story.append(Paragraph(
        f"{_lab('period', korean_ok)}: {summary.period} &nbsp;&nbsp; "
        f"{_lab('generated', korean_ok)}: {now}", body))
    story.append(Spacer(1, 8 * mm))

    # KPI 요약 표
    story.append(Paragraph(_lab("kpi", korean_ok), h2))
    kpi_rows = [
        [_lab("total_inspected", korean_ok), f"{summary.total_inspected}"],
        [_lab("defect_count", korean_ok), f"{summary.defect_count}"],
        [_lab("process_defect_ppm", korean_ok), f"{summary.process_defect_ppm:.3f}"],
        [_lab("inspection_defect_rate_pct", korean_ok),
         f"{summary.inspection_defect_rate_pct:.3f}"],
        [_lab("auto_inspection_rate_pct", korean_ok),
         f"{summary.auto_inspection_rate_pct:.3f}"],
        [_lab("storage_mes_rate_pct", korean_ok),
         f"{summary.storage_mes_rate_pct:.3f}"],
        [_lab("avg_proc_time_ms", korean_ok),
         ("-" if summary.avg_proc_time_ms is None else f"{summary.avg_proc_time_ms:.2f}")],
    ]
    t = Table(kpi_rows, colWidths=[90 * mm, 70 * mm])
    t.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, -1), font),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("BACKGROUND", (0, 0), (0, -1), colors.whitesmoke),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("FONTSIZE", (0, 0), (-1, -1), 10),
    ]))
    story.append(t)
    story.append(Spacer(1, 6 * mm))

    # 처리속도 분포(p50/p95/p99) — FAT 기준 300ms/ea 는 꼬리까지 봐야 한다.
    pct = proc_time_percentiles(rows)
    story.append(Paragraph(_lab("proc_pct", korean_ok), h2))
    ptab = Table(
        [[_lab("p50", korean_ok), _lab("p95", korean_ok), _lab("p99", korean_ok)],
         [("-" if pct["p50"] is None else f"{pct['p50']:.0f}"),
          ("-" if pct["p95"] is None else f"{pct['p95']:.0f}"),
          ("-" if pct["p99"] is None else f"{pct['p99']:.0f}")]],
        colWidths=[53 * mm, 53 * mm, 54 * mm],
    )
    ptab.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, -1), font),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("FONTSIZE", (0, 0), (-1, -1), 10),
    ]))
    story.append(ptab)
    story.append(Spacer(1, 6 * mm))

    # KPI 목표 대비 달성 판정(§1.1/§1.2) — 인수 심사용.
    # 달성/미달은 색만이 아니라 기호(✓/✗)+문자를 함께 표기한다(색각 이상 고려).
    story.append(Paragraph(_lab("targets", korean_ok), h2))
    tgt_rows = [[_lab("kpi_item", korean_ok), _lab("target", korean_ok),
                 _lab("actual", korean_ok), _lab("achieved", korean_ok)]]
    tgt_style = [
        ("FONTNAME", (0, 0), (-1, -1), font),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
        ("FONTSIZE", (0, 0), (-1, -1), 10),
        ("ALIGN", (1, 1), (-1, -1), "CENTER"),
    ]
    for i, (key, ko, latin, target_text, passed) in enumerate(
        evaluate_targets(summary, rows), start=1
    ):
        # 기호는 ASCII 로만 쓴다: 한글 CID 폰트에 ✓/✗(U+2713/2717) 글리프가 없어
        # 조용히 빈칸으로 렌더링돼 "색 단독 표기"가 되어버린다(색각 이상 위험).
        if passed is None:
            mark, ink = f"- {_lab('na', korean_ok)}", AXIS_INK
        elif passed:
            mark, ink = f"O {_lab('pass_ko', korean_ok)}", STATUS_GOOD
        else:
            mark, ink = f"X {_lab('fail_ko', korean_ok)}", STATUS_CRITICAL
        tgt_rows.append([
            ko if korean_ok else latin,
            target_text if korean_ok else target_text.replace(" 이하", " max"),
            target_actual_text(key, summary, rows),
            mark,
        ])
        tgt_style.append(("TEXTCOLOR", (3, i), (3, i), colors.HexColor(ink)))
    tt = Table(tgt_rows, colWidths=[58 * mm, 34 * mm, 34 * mm, 34 * mm])
    tt.setStyle(TableStyle(tgt_style))
    story.append(tt)
    story.append(Spacer(1, 8 * mm))

    # 일자별 공정불량률 추세(차트) — 표만으로는 추세가 안 보인다(M12 시각화).
    story.append(Paragraph(_lab("trend", korean_ok), h2))
    story.append(_daily_trend_chart(aggregate_daily(rows), font, korean_ok))
    story.append(Spacer(1, 6 * mm))

    # 불량유형별 집계
    story.append(Paragraph(_lab("defect_breakdown", korean_ok), h2))
    breakdown = aggregate_defects(rows)
    story.append(_defect_bar_chart(breakdown, font, korean_ok))
    story.append(Spacer(1, 4 * mm))
    if breakdown:
        data = [[_lab("code", korean_ok), _lab("count", korean_ok)]]
        data += [[_defect_label(c, korean_ok), str(n)] for c, n in breakdown]
    else:
        data = [[_lab("code", korean_ok), _lab("count", korean_ok)],
                [_lab("none", korean_ok), "0"]]
    dt = Table(data, colWidths=[90 * mm, 70 * mm])
    dt.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, -1), font),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
        ("FONTSIZE", (0, 0), (-1, -1), 10),
    ]))
    story.append(dt)
    story.append(Spacer(1, 8 * mm))

    # 일자별 표
    story.append(Paragraph(_lab("daily", korean_ok), h2))
    daily = aggregate_daily(rows)
    ddata = [[_lab("date", korean_ok), _lab("inspected", korean_ok),
              _lab("defects", korean_ok)]]
    if daily:
        ddata += [[d, str(i), str(x)] for d, i, x in daily]
    else:
        ddata += [[_lab("none", korean_ok), "0", "0"]]
    ddt = Table(ddata, colWidths=[70 * mm, 45 * mm, 45 * mm])
    ddt.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, -1), font),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
    ]))
    story.append(ddt)

    doc.build(story)
    return buf.getvalue()


# ---- XLSX -----------------------------------------------------------------


def render_xlsx(summary: KpiSummary, rows: list[Inspection]) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    bold = Font(bold=True)
    head_fill = PatternFill("solid", fgColor="DDDDDD")

    # 시트 1: KPI 요약 (한글 라벨)
    ws = wb.active
    ws.title = "KPI"
    ws["A1"] = LABELS["title"][0]
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = LABELS["period"][0]
    ws["B2"] = summary.period
    ws["A3"] = LABELS["generated"][0]
    ws["B3"] = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")

    r = 5
    kpi_pairs = [
        (LABELS["total_inspected"][0], summary.total_inspected),
        (LABELS["defect_count"][0], summary.defect_count),
        (LABELS["process_defect_ppm"][0], summary.process_defect_ppm),
        (LABELS["inspection_defect_rate_pct"][0], summary.inspection_defect_rate_pct),
        (LABELS["auto_inspection_rate_pct"][0], summary.auto_inspection_rate_pct),
        (LABELS["storage_mes_rate_pct"][0], summary.storage_mes_rate_pct),
        (LABELS["avg_proc_time_ms"][0], summary.avg_proc_time_ms),
    ]
    ws.cell(r, 1, LABELS["kpi"][0]).font = bold
    r += 1
    for label, value in kpi_pairs:
        ws.cell(r, 1, label)
        ws.cell(r, 2, value)
        r += 1

    # 처리속도 백분위(FAT 기준 300ms/ea 는 꼬리까지 확인해야 한다).
    pct = proc_time_percentiles(rows)
    r += 1
    ws.cell(r, 1, LABELS["proc_pct"][0]).font = bold
    r += 1
    for key in ("p50", "p95", "p99"):
        ws.cell(r, 1, LABELS[key][0])
        ws.cell(r, 2, pct[key])
        r += 1

    # KPI 목표 대비 달성(인수 심사용) — PDF 와 동일 기준.
    r += 1
    ws.cell(r, 1, LABELS["targets"][0]).font = bold
    r += 1
    for col, key in enumerate(
        ("kpi_item", "target", "actual", "achieved"), start=1
    ):
        c = ws.cell(r, col, LABELS[key][0])
        c.font = bold
        c.fill = head_fill
    r += 1
    for key, ko, _latin, target_text, passed in evaluate_targets(summary, rows):
        ws.cell(r, 1, ko)
        ws.cell(r, 2, target_text)
        ws.cell(r, 3, target_actual_text(key, summary, rows))
        ws.cell(
            r, 4,
            LABELS["na"][0] if passed is None
            else (LABELS["pass_ko"][0] if passed else LABELS["fail_ko"][0]),
        )
        r += 1

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 12

    # 시트 2: 불량유형별 집계 (시트명에 / 등 금지문자 사용 불가 -> 안전한 라틴명)
    ws2 = wb.create_sheet("Defects")
    ws2.cell(1, 1, LABELS["code"][0]).font = bold
    ws2.cell(1, 2, LABELS["count"][0]).font = bold
    ws2["A1"].fill = head_fill
    ws2["B1"].fill = head_fill
    rr = 2
    for code, n in aggregate_defects(rows):
        ws2.cell(rr, 1, _defect_label(code, True))
        ws2.cell(rr, 2, n)
        rr += 1
    ws2.column_dimensions["A"].width = 22

    # 시트 3: 일자별 검사/불량 (시트명은 안전한 라틴명)
    ws3 = wb.create_sheet("Daily")
    for ci, key in enumerate(("date", "inspected", "defects"), start=1):
        c = ws3.cell(1, ci, LABELS[key][0])
        c.font = bold
        c.fill = head_fill
    rr = 2
    for day, ins, dfx in aggregate_daily(rows):
        ws3.cell(rr, 1, day)
        ws3.cell(rr, 2, ins)
        ws3.cell(rr, 3, dfx)
        rr += 1
    ws3.column_dimensions["A"].width = 14

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
